#!/usr/bin/env python3
"""
build-fix-tracker.py

Build a multi-tab .xlsx fix tracker from a JSON findings file. Used by
prompts/triage-final-check.md to turn a final-check report into a shareable,
tab-per-fix-type tracker. Drag the .xlsx into Google Drive to open it as a
Google Sheet with the tabs intact.

Why a file, not a direct Google Sheet: the Drive connector only takes inline
content, and a binary .xlsx cannot be pushed through it reliably, so we build
the exact file and import it with one drag.

Input JSON:
{
  "title": "Client - Pre-launch fix tracker (from final-check YYYY-MM-DD)",
  "overview": ["line one of context", "line two", ...],   // optional
  "tabs": {
    "Content":    [["Page","URL","Section","Issue","Fix","Owner","Severity"], ...rows],
    "Compliance": [...],
    ...
  }
}
Each tab is a list of rows; row 1 may be the header (7 cells) or you can omit it
and the standard header is added. Owner is "AI" or "Human" (or "AI / Human").

Usage:
  python .claude/tools/build-fix-tracker.py findings.json build-log/<client>-prelaunch-fixes.xlsx
"""
import json, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLS = ["Page", "URL", "Section", "Issue", "Fix", "Owner", "Severity"]
WIDTHS = [16, 34, 20, 46, 54, 16, 12]


def main():
    if len(sys.argv) < 3:
        sys.exit("usage: build-fix-tracker.py findings.json output.xlsx")
    spec = json.load(open(sys.argv[1], encoding="utf-8"))
    out = sys.argv[2]

    wb = openpyxl.Workbook()
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="262B44")
    wrap = Alignment(wrap_text=True, vertical="top")

    ov = wb.active
    ov.title = "Overview"
    ov.column_dimensions["A"].width = 115
    ov.append([spec.get("title", "Pre-launch fix tracker")])
    ov["A1"].font = Font(bold=True, size=14)
    for line in spec.get("overview", []):
        ov.append([line])

    for name, rows in spec.get("tabs", {}).items():
        ws = wb.create_sheet(name[:31])
        # add standard header unless the first row already looks like one
        if not rows or [str(c).strip() for c in rows[0][:2]] != COLS[:2]:
            ws.append(COLS)
        for r in rows:
            ws.append(r)
        for c in range(1, len(COLS) + 1):
            ws.cell(1, c).font = head
            ws.cell(1, c).fill = fill
        for i, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        ws.freeze_panes = "A2"

    wb.save(out)
    n = sum(ws.max_row - 1 for ws in wb.worksheets if ws.title != "Overview")
    print(f"wrote {out}: {len(wb.worksheets)} tabs, {n} rows")
    print("tabs:", [ws.title for ws in wb.worksheets])
    print("To share as a Google Sheet: drag this .xlsx into Google Drive.")


if __name__ == "__main__":
    main()
